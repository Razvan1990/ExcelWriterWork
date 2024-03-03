import openpyxl
from openpyxl.styles import Side, PatternFill, Alignment, Font
import os
import shutil

EXCEL_FILE_NAME = "TB__FAS__8_2407_2__G05-1830_2024-02-02-11-39-49_backup_orig_testing.xlsx"
FOLDER_NAME_GENERATION = "generations"
NEEDED_SHEET = "Details"
INTRODUCTION_SHEET = "Deckblatt"
MISRA_ARGUMENTATION_WORD = "siehe 3"
TICKET_COVERAGE = "FASACHT-12345"


def get_dictionary_of_colours(excel_file=EXCEL_FILE_NAME, sheet_name=INTRODUCTION_SHEET):
    '''
    we will traverse through the first sheet and take the colors from the
    legend sheets

    :param excel_file:
    :param sheet_name:
    :return: a dictionary of colors that we need
    '''
    dict_colors = dict()
    # need colors from F31, F29, F33
    workbook = openpyxl.load_workbook(excel_file)
    sheet_active = workbook[sheet_name]
    red_color = sheet_active.cell(31, 6).fill.fgColor.rgb
    green_color = sheet_active.cell(29, 6).fill.fgColor.rgb
    dirty_green_color = sheet_active.cell(33, 6).fill.fgColor.rgb
    normal_white_color = "00FFFFFF"
    dict_colors.update({"WHITE": normal_white_color})
    dict_colors.update({"RED": red_color})
    dict_colors.update({"GREEN": green_color})
    dict_colors.update({"DIRTY_GREEN": dirty_green_color})
    return dict_colors


def write_to_excel(excel_file=EXCEL_FILE_NAME, generation_folder=FOLDER_NAME_GENERATION, sheet_name=NEEDED_SHEET,
                   test_coverage_ticket=TICKET_COVERAGE):
    try:
        string_report_misra = ""
        string_report_coverage = ""
        dict_colours = get_dictionary_of_colours()
        workbook = openpyxl.load_workbook(excel_file)
        sheet_active = workbook[sheet_name]
        # first lets iterate over all rows of AH
        number_rows = len(sheet_active["AH"]) - 1
        print(number_rows)
        column_needed_misra = ["AH"]
        for column in column_needed_misra:
            for row in range(1, number_rows):
                if sheet_active["{}{}".format(column, row)].value is None:
                    continue
                # put condition to check if it is NA and has the fill red
                elif sheet_active["{}{}".format(column, row)].value == "NA" and sheet_active[
                    "{}{}".format(column, row)].fill.fgColor.rgb == dict_colours["RED"]:
                    sheet_active["{}{}".format(column, row)].value = MISRA_ARGUMENTATION_WORD
                    sheet_active["{}{}".format(column, row)].alignment = Alignment(horizontal="center", wrap_text=True)
                    sheet_active["{}{}".format(column, row)].font = Font(name="Calibri", size=10)
                    sheet_active["{}{}".format(column, row)].fill = PatternFill(fill_type="solid",
                                                                                start_color=dict_colours["WHITE"],
                                                                                end_color=dict_colours["WHITE"])
                    string_report_misra += "Cell {}{} has been completed with {}\n".format(column, row,
                                                                                           MISRA_ARGUMENTATION_WORD)
        workbook.save(excel_file)
        # write to report
        if not os.path.exists(generation_folder):
            os.mkdir(generation_folder)
        with open(file="report_misra.txt", mode="w", encoding="utf-8") as file:
            file.write(string_report_misra)
        # move to generations
        shutil.move(os.path.join(os.getcwd(), "report_misra.txt"),
                    os.path.join(os.getcwd(), generation_folder, "report_misra.txt"))

        '''
        now we will put the ticket where we have no 100% percent coverage in column N
        in fact as some percentages there are not 100 % always and are green, we just check to see if we have the red color on the cell
        '''
        columns_needed_argumentations = ["N", "O", "Q"]
        for column in columns_needed_argumentations[:-1]:
            for row in range(1, number_rows):
                if sheet_active["{}{}".format(column, row)].value is None:
                    continue
                elif sheet_active["{}{}".format(column, row)].fill.fgColor.rgb == dict_colours["RED"]:
                    sheet_active["{}{}".format(columns_needed_argumentations[2], row)].value = test_coverage_ticket
                    sheet_active["{}{}".format(columns_needed_argumentations[2], row)].alignment = Alignment(
                        horizontal="center", wrap_text=True)
                    sheet_active["{}{}".format(columns_needed_argumentations[2], row)].font = Font(name="Calibri",
                                                                                                   size=10)
                    sheet_active["{}{}".format(columns_needed_argumentations[2], row)].fill = PatternFill(
                        fill_type="solid", start_color=dict_colours["WHITE"], end_color=dict_colours["WHITE"])
                    string_report_coverage += "Cell {}{} has been completed with {}\n".format(
                        columns_needed_argumentations[2], row, TICKET_COVERAGE)
        workbook.save(excel_file)
        # write report to coverage
        with open(file="report_coverage.txt", mode="w", encoding="utf-8") as file:
            file.write(string_report_coverage)
        shutil.move(os.path.join(os.getcwd(), "report_coverage.txt"),
                    os.path.join(os.getcwd(), generation_folder, "report_coverage.txt"))
    except:
        raise Exception("Close the excel file!")


def colour_excel(excel_file=EXCEL_FILE_NAME, sheet_name=NEEDED_SHEET):
    '''
    we will just iterate through all the file and just check where we have red fill
    :param excel_file:
    :param sheet_name:
    :return: the coloured excel file
    '''
    try:
        dict_colours = get_dictionary_of_colours()
        workbook = openpyxl.load_workbook(excel_file)
        sheet_active = workbook[sheet_name]
        number_max_columns = sheet_active.max_column - 1  # because last column is empty
        number_max_rows = sheet_active.max_row - 1
        for index_column in range(1, number_max_columns):
            for index_row in range(1, number_max_rows):
                if sheet_active.cell(index_row, index_column).fill.fgColor.rgb == dict_colours["RED"]:
                    sheet_active.cell(index_row, index_column).fill = PatternFill(fill_type="solid",
                                                                                  start_color=dict_colours["DIRTY_GREEN"],
                                                                                  end_color=dict_colours["DIRTY_GREEN"])
                    #make everything to be write in black
                    sheet_active.cell(index_row, index_column). font = Font(name="Calibri", size=10, color="FF000000")
        workbook.save(excel_file)
    except:
        raise Exception("Please close the excel file!")


if __name__ == "__main__":
    # print(get_dictionary_of_colours())
    write_to_excel()
    colour_excel()
